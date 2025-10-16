# Copyright (c) 2025 Bytedance Ltd. and/or its affiliates
# SPDX-License-Identifier: GPL-3.0-or-later
import os
import traceback
from openpyxl import Workbook
from openpyxl.styles import Alignment
from ..common.homeless import timing_decorator


# 一个实例对应一个excel
class ExcelUtil:
    wb: Workbook

    def __init__(self):
        self.wb = Workbook()

    def updateData(self, data: list):
        ws = self.wb.active

        for row in data:
            ws.append(row)

    @timing_decorator
    def saveExcel(self, dir: str = None, file_name: str = "result.xlsx"):
        try:
            if dir is None:
                self.wb.save(file_name)
            else:
                full_path = os.path.join(dir, file_name)
                # save当前工作目录
                self.wb.save(full_path)
        except Exception as e:
            traceback.print_exc()

    def mergeCeil(self, start_row: int, start_col: int, end_row: int, end_col: int):
        ws = self.wb.active

        # 获取指定单元格
        start_cell = ws.cell(row=start_row, column=start_col)
        end_cell = ws.cell(row=end_row, column=end_col)

        # 获取单元格的名称
        start_cell_name = start_cell.coordinate
        end_cell_name = end_cell.coordinate

        # 合并单元格
        ws.merge_cells(f"{start_cell_name}:{end_cell_name}")

    def getCeilName(self, row: int, column: int):
        ws = self.wb.active

        current_ceil = ws.cell(row=row, column=column)

        cell_name = current_ceil.coordinate

        return f"{cell_name}"

    # 将对传入的行、列配置支持换行
    def alignmentCell(self, row_len: int, col_len: int):
        ws = self.wb.active

        for i in range(1, row_len):
            for j in range(1, col_len):
                cell = ws.cell(row=i, column=j)

                # 使单元格中的文本自动换行
                cell.alignment = Alignment(wrap_text=True)

    # 设置excel中的单元格自适应宽度
    def setColumnAutoWidth(self, max_width: int = 0):
        ws = self.wb.active
        # 设置列宽
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 10
            if max_width != 0 and adjusted_width >= max_width:
                adjusted_width = max_width

            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    def setRowHeight(self, height: int):
        ws = self.wb.active
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = height


# data = [
#     ['Name', 'Age', 'Gender'],
#     ['John', 30, 'Male'],
#     ['Lisa', 28, 'Female'],
#     ['Linda', 32, 'Female'],
#     ['Michael', 41, 'Male']
# ]
# demo = ExcelUtil()
# demo.updateData(data)
# demo.saveExcel()
